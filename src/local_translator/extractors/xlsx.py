from __future__ import annotations

from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from local_translator.extractors.base import BaseExtractor, ExtractedDocument


class XlsxExtractor(BaseExtractor):
    suffixes = (".xlsx",)

    _TEXT_CELL_REFS_KEY: Final[str] = "text_cell_refs"

    @staticmethod
    def _is_translatable_text_cell(cell: Cell) -> bool:
        value = cell.value
        if not isinstance(value, str):
            return False

        if cell.data_type == "f":
            return False

        if value.strip() == "":
            return False

        return True

    @staticmethod
    def _preserve_edge_whitespace(original: str, translated: str) -> str:
        leading_len = len(original) - len(original.lstrip())
        trailing_len = len(original) - len(original.rstrip())
        if leading_len == 0 and trailing_len == 0:
            return translated

        leading_ws = original[:leading_len]
        trailing_ws = original[len(original) - trailing_len :] if trailing_len else ""
        return f"{leading_ws}{translated.strip()}{trailing_ws}"

    def extract(self, file_path: Path) -> ExtractedDocument:
        workbook = load_workbook(str(file_path))
        segments: list[str] = []
        text_cell_refs: list[tuple[str, str]] = []

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                for cell in row:
                    if self._is_translatable_text_cell(cell):
                        assert isinstance(cell.value, str)
                        segments.append(cell.value)
                        text_cell_refs.append((worksheet.title, cell.coordinate))

        return ExtractedDocument(
            file_path=file_path,
            segments=segments,
            metadata={self._TEXT_CELL_REFS_KEY: text_cell_refs},
        )

    def reconstruct(self, extracted: ExtractedDocument, translated_segments: list[str], output_path: Path) -> None:
        text_cell_refs = extracted.metadata.get(self._TEXT_CELL_REFS_KEY)
        if not isinstance(text_cell_refs, list):
            raise ValueError("Missing XLSX text cell references in extracted metadata.")

        if len(text_cell_refs) != len(translated_segments):
            raise ValueError(
                "Mismatch between extracted text cells and translated segments "
                f"({len(text_cell_refs)} != {len(translated_segments)})."
            )

        workbook = load_workbook(str(extracted.file_path))

        for (sheet_name, coordinate), translated_text in zip(text_cell_refs, translated_segments, strict=True):
            worksheet = workbook[sheet_name]
            cell = worksheet[coordinate]
            if self._is_translatable_text_cell(cell):
                assert isinstance(cell.value, str)
                cell.value = self._preserve_edge_whitespace(cell.value, translated_text)

        workbook.save(str(output_path))
